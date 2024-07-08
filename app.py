# OpenCV library for image processing and computer vision tasks.
import cv2

# Operating system interface for directory operations and file handling.
import os

# Web framework for building web applications in Python.
from flask import Flask, request, render_template

# Classes for working with dates and times.
from datetime import date, datetime

# Library for numerical computing, used for handling arrays and matrices.
import numpy as np

# Classifier for k-nearest neighbors algorithm from scikit-learn.
from sklearn.neighbors import KNeighborsClassifier

# Library for data manipulation and analysis.
import pandas as pd

# Library for saving and loading Python objects (used here to save and load machine learning models).
import joblib

# Library for reading and writing Excel files.
from openpyxl import Workbook, load_workbook

class FlaskAppSetup:
    def __init__(self):

        # Initializes a Flask application.
        self.app = Flask(__name__)

        '''
        Initializes a boolean variable app_running and sets it to True. 
        This variable is used later to control the running state of the Flask application.
        '''

        self.app_running = True  

class ModelTrainer:
    def __init__(self):
        pass

    def train_model(self):
        '''
        Initializes empty lists faces and labels to store the face images and their corresponding 
        labels (user names).
        '''
        faces = []
        labels = []

        '''
        Retrieves the list of directories (user folders) within the 'static/faces' 
        directory using os.listdir().
        '''
        userlist = os.listdir('static/faces')

        # Iterates over each user directory in userlist.
        for user in userlist:

            # Iterates over each image file (imgname) within the current user's directory.
            for imgname in os.listdir(f'static/faces/{user}'):

                # Reads each image file (img) using OpenCV's cv2.imread() function.
                img = cv2.imread(f'static/faces/{user}/{imgname}')

                # Resizes the image to a fixed size of 50x50 pixels using cv2.resize() and stores it in resized_face.
                resized_face = cv2.resize(img, (50, 50))

                # Flattens the resized face image into a 1D array using .ravel() and appends it to the faces list.
                faces.append(resized_face.ravel())

                # Appends the user's name (user) to the labels list, representing the label for the corresponding face image.
                labels.append(user)

        # Converts the faces list into a NumPy array for compatibility with scikit-learn.
        faces = np.array(faces)

        # Initializes a K-Nearest Neighbors classifier (knn) with 5 neighbors.
        knn = KNeighborsClassifier(n_neighbors=5)

        '''
        Trains the KNN classifier (knn) using the face images (faces) as training data and their 
        corresponding labels (labels).
        '''
        knn.fit(faces, labels)

        '''
        Saves the trained KNN model (knn) to a file named 'static/face_recognition_model.pkl' 
        using joblib's joblib.dump() function.
        '''
        joblib.dump(knn, 'static/face_recognition_model.pkl')


class ImageProcessing:
    def __init__(self):

        # Number of images to capture for each user during registration (initializes its value to 30).
        self._nimgs = 30  # Protected attribute

        # Background image loaded from "bg.png" file using OpenCV library’s function cv2.imread.
        self.imgBackground = cv2.imread("bg.png")  # Protected attribute

        # Cascade classifier for face detection.
        self._face_detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')  # Protected attribute
        self.model_trainer = ModelTrainer()


    def extract_faces(self, img):
        '''
        Defines a function extract_faces(img) that takes an image (img) as input attempts to detect faces 
        in the image using the face_detector cascade classifier. It converts the image to grayscale, 
        detects faces using detectMultiScale() function.
        '''
        try:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            face_points = self._face_detector.detectMultiScale(gray, 1.2, 5, minSize=(20, 20))
            # # returns the coordinates of the detected faces.
            return face_points
        except:
            # If no faces are detected or an exception occurs, it returns an empty list.
            return []
        
        
        
    def identify_face(self, facearray):
        """
        Defines a function identify_face(facearray) that takes an array of face images (facearray) as input.
        """

        '''
        Identifies the faces using a pre-trained face recognition model loaded from the file 
        'static/face_recognition_model.pkl'.
        '''
        model = joblib.load('static/face_recognition_model.pkl')

        # It returns the predicted labels for the input face images.
        return model.predict(facearray)



# Current date formatted as "dd-Month-YYYY".
datetoday2 = date.today().strftime("%d-%B-%Y")


class ParentSetupFiles:
    def __init__(self, datetoday):
        self.datetoday = datetoday


class SetupFiles(ParentSetupFiles):
    def __init__(self, datetoday):
        # Call the constructor of the parent class
        super().__init__(datetoday)

    def setup_files(self):
        # Implement file setup logic here
        '''
        Checks if the attendance CSV file for the current date exists in the 'Attendance' directory. 
        If it doesn't exist, it creates a new CSV file named 'Attendance-{datetoday}.csv' with headers 
        'Name', 'Roll', and 'Time'.
        '''

        if f'Attendance-{self.datetoday}.csv' not in os.listdir('Attendance'):
            with open(f'Attendance/Attendance-{self.datetoday}.csv', 'w') as f:
                f.write('Name,Roll,Time')



class SetupFolders:
    def __init__(self):
        pass

    def setup(self):
        # Implement folder setup logic here

        '''
        Checks if the directory named 'Attendance', 'static', 'static/faces' exists. 
        If it doesn't exist, it creates the directory.
        'static/faces' directory is used to store images of faces
        '''
        for folder in ['Attendance', 'static', 'static/faces']:
            if not os.path.isdir(folder):
                os.makedirs(folder)
        pass



class FolderFileSetup(SetupFolders,SetupFiles):
    def __init__(self):
        datetoday2 = date.today().strftime("%d-%B-%Y")
        # super().__init__()  # Call the constructor of the parent class (SetupFolders)
        # Current date formatted as "mm_dd_yy".
        # Call the constructors of both parent classes with appropriate arguments
        SetupFolders.__init__(self)
        SetupFiles.__init__(self, self.datetoday2)

        self.datetoday = date.today().strftime("%m_%d_%y")  # Protected attribute

        # Current date formatted as "dd-Month-YYYY".
        self.datetoday2 = date.today().strftime("%d-%B-%Y")  # Protected attribute

    def setup_folders(self):
        # Call SetupFolders class to handle folder setup logic
        # folder_setup = SetupFolders()
        # folder_setup.setup()
        # Call the setup method of the parent class (SetupFolders)
        self.setup()


    def setup_files(self):
        # Implement file setup logic here
        file_setup = SetupFiles(self.datetoday)
        file_setup.setup_files()

from flask import render_template



class ModelTrainer:
    def __init__(self):
        pass

    def train_model(self):
        faces = []
        labels = []
        userlist = os.listdir('static/faces')
        for user in userlist:
            for imgname in os.listdir(f'static/faces/{user}'):
                img = cv2.imread(f'static/faces/{user}/{imgname}')
                resized_face = cv2.resize(img, (50, 50))
                faces.append(resized_face.ravel())
                labels.append(user)
        faces = np.array(faces)
        knn = KNeighborsClassifier(n_neighbors=5)
        knn.fit(faces, labels)
        joblib.dump(knn, 'static/face_recognition_model.pkl')







class AttendanceMarker:
    def __init__(self, app_running, imgBackground, extract_attendance, totalreg, datetoday2, extract_faces, identify_face, add_attendance):
        self.app_running = app_running
        self.imgBackground = imgBackground
        self.extract_attendance = extract_attendance
        self.totalreg = totalreg
        self.datetoday2 = datetoday2
        self.extract_faces = extract_faces
        self.identify_face = identify_face
        self.add_attendance = add_attendance

    def mark_attendance(self):

        '''
        Calls the extract_attendance() function to retrieve attendance data. It retrieves 
        lists of names, rolls, times, and the length of the attendance records.
        '''
        names, rolls, times, l = self.extract_attendance()

        '''
        Checks if the trained face recognition model ('face_recognition_model.pkl') is present in 
        the 'static' folder. If the model is not found, it indicates that face recognition cannot be 
        performed.
        '''
        if 'face_recognition_model.pkl' not in os.listdir('static'):
            '''
            If the trained model is not found, it renders the 'attendance.html' template with the provided 
            message indicating that the model is missing. It also passes attendance data (names, rolls, 
            times, l), total registered users count (totalreg()), current date (datetoday2), and the 
            message (mess) to the template.
            '''
            return render_template('attendance.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.totalreg(), datetoday2=self.datetoday2, mess='There is no trained model in the static folder. Please add a new face to continue.')

        
        # Starts capturing video from the default camera (index 0) using OpenCV. 
        ret = True
        cap = cv2.VideoCapture(0)
        
        while ret and self.app_running:
            '''
            Enters a loop to continuously capture frames while ret is true (meaning frames are 
            being successfully read) and app_running is true (indicating the application is still running).
            '''
            ret, frame = cap.read()
            print(len(self.extract_faces(frame)))

            # Checks if faces are detected in the current frame.
            if len(self.extract_faces(frame)) > 0:
                '''
                If faces are detected, it extracts the coordinates of the first detected face and 
                proceeds with face recognition.
                '''
                (x, y, w, h) = self.extract_faces(frame)[0]

                '''
                    • These lines draw rectangles around the detected face:
                    ◦ The first rectangle outlines the face area with a thin border.
                    ◦ The second rectangle highlights the area above the face (presumably for
                      displaying text) by drawing a filled rectangle.
                '''
                cv2.rectangle(frame, (x, y), (x + w, y + h), (86, 32, 251), 1)
                cv2.rectangle(frame, (x, y), (x + w, y - 40), (86, 32, 251), -1)

                '''
                These lines extract the detected face region from the frame, resize it to a standard 
                size (50x50 pixels).
                '''
                face = cv2.resize(frame[y:y + h, x:x + w], (50, 50))

                '''
                It then pass it to a function identify_face() for recognition. 
                The recognized person's name is stored in identified_person.
                '''
                identified_person = self.identify_face(face.reshape(1, -1))[0]
                print("Identified person:", identified_person)

                # updates the attendance record with the identified person's name.
                self.add_attendance(identified_person)

                # These lines draw rectangles.
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 1)
                cv2.rectangle(frame, (x, y), (x + w, y + h), (50, 50, 255), 2)
                cv2.rectangle(frame, (x, y - 40), (x + w, y), (50, 50, 255), -1)

                # adds text displaying the identified person's name above the face.
                cv2.putText(frame, f'{identified_person}', (x, y - 15), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)
                cv2.rectangle(frame, (x, y), (x + w, y + h), (50, 50, 255), 1)
            
            # Calculate the center position for the webcam window
            center_x = (self.imgBackground.shape[1] - frame.shape[1]) // 2
            center_y = (self.imgBackground.shape[0] - frame.shape[0]) // 2

            '''
            This section places the processed frame (with rectangles and text) onto a background frame (imgBackground) 
            to create a composite image. This composite image is then displayed in the GUI window.
            '''
            self.imgBackground[center_y:center_y + frame.shape[0], center_x:center_x + frame.shape[1]] = frame

            # Displays the current frame with the detected faces and identification information.
            cv2.imshow('Attendance', self.imgBackground)

            '''
            Checks if the 'Esc' key is pressed or if app_running is set to False. If either condition is met, 
            it sets app_running to False and breaks out of the loop.
            '''
            if cv2.waitKey(1) == 27 or not self.app_running:
                self.app_running = False
                break
        
        # Releases the video capture object (cap)
        cap.release()

        # closes all OpenCV windows.
        cv2.destroyAllWindows()

        '''
        After exiting the loop, it renders the 'attendance.html' template with the attendance data (names, rolls, times, l), 
        total registered users count (totalreg()), and current date (datetoday2).
        '''
        names, rolls, times, l = self.extract_attendance()
        
        return render_template('attendance.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.totalreg(), datetoday2=self.datetoday2)


class FaceExtractor:
    def __init__(self, face_detector):
        self.face_detector = face_detector

    def extract_faces(self, img):
        '''
        Defines a function extract_faces(img) that takes an image (img) as input attempts to detect faces 
        in the image using the face_detector cascade classifier. It converts the image to grayscale, 
        detects faces using detectMultiScale() function.
        '''
        try:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            face_points = self._face_detector.detectMultiScale(gray, 1.2, 5, minSize=(20, 20))
            return face_points
            # returns the coordinates of the detected faces.
        except:
            # If no faces are detected or an exception occurs, it returns an empty list.
            return []
        


class AttendanceManager(ModelTrainer):
    #--------------------------write logic------------------------------------------------
    def __init__(self, image_processor, folder_file_setup):
        super().__init__()  # Call the constructor of the parent class (ModelTrainer)
        # Initialize the AttendanceManager class
        self.image_processor = image_processor
        self.folder_file_setup = folder_file_setup
        self._face_detector = self.image_processor._face_detector  # Assuming face_detector is defined in ImageProcessing
        if self._face_detector.empty():
            print("Cascade classifier not loaded properly.")
        # self.model_trainer = ModelTrainer()

    def totalreg(self):
        '''
        Defines a function totalreg() that returns the total number of registered faces by counting the 
        number of files in the 'static/faces' directory.
        '''
        return len(os.listdir('static/faces'))
    
    def __init__(self):

        # Number of images to capture for each user during registration (initializes its value to 10).
        self._nimgs = 30  # Protected attribute

        # Background image loaded from "bg.png" file using OpenCV library’s function cv2.imread.
        self.imgBackground = cv2.imread("bg.png")  # Protected attribute

        # Cascade classifier for face detection.
        self._face_detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')  # Protected attribute



    def extract_faces(self, img):

        
        self.face_extractor = FaceExtractor(self._face_detector)

        
        
    def identify_face(self, facearray):
        """
        Defines a function identify_face(facearray) that takes an array of face images (facearray) as input.
        """

        '''
        Identifies the faces using a pre-trained face recognition model loaded from the file 
        'static/face_recognition_model.pkl'.
        '''
        model = joblib.load('static/face_recognition_model.pkl')

        # It returns the predicted labels for the input face images.
        return model.predict(facearray)
    
    def train_model(self):
        self.model_trainer.train_model()
    
    def extract_attendance(self):
        '''
        Extracts attendance data from a CSV file.

        Returns:
            names (pandas.Series): Series containing names of attendees.
            rolls (pandas.Series): Series containing roll numbers of attendees.
            times (pandas.Series): Series containing timestamp of attendance.
            l (int): Length of the attendance data.
        '''

        # Current date formatted as "mm_dd_yy".
        datetoday4 = date.today().strftime("%m_%d_%y")

        '''
        This line reads the CSV file containing attendance data for the current date (datetoday). 
        It uses pandas' read_csv() function to read the CSV file into a DataFrame named df. 
        The file path is formatted dynamically to include the current date in the filename.
        '''
        df = pd.read_csv(f'Attendance/Attendance-{datetoday4}.csv')

        '''
        These lines extract the 'Name', 'Roll', and 'Time' columns from the DataFrame df and assign them 
        to variables names, rolls, and times, respectively. Each variable holds a Series containing the 
        data from the corresponding column.
        '''
        names = df['Name']
        rolls = df['Roll']
        times = df['Time']


        '''
        This line calculates the length of the DataFrame df using the len() function and assigns it 
        to the variable l. It represents the number of rows (entries) in the DataFrame, which corresponds 
        to the number of attendance records for the current date.
        '''
        l = len(df)

        '''
        This line returns a tuple containing the extracted 'Name', 'Roll', 'Time', and the length of the 
        DataFrame (names, rolls, times, l). This allows the caller of the function to access this data 
        for further processing or display.
        '''
        return   names, rolls, times, l
    
    def add_attendance(self, name):
        '''
        Add attendance record to CSV file.

        Parameters:
            name (str): Name in the format 'username_userid'.

        Returns:
            None

        Raises:
            ValueError: If the provided userid is not a valid integer.
        '''

        '''
        These lines split the name string using the underscore ('_') as a delimiter. 
        It assumes that the name string follows the format "username_userid". The first part before 
        the underscore is extracted as username, and the second part after the underscore is extracted 
        as userid.
        '''
        # Placeholder implementation
        username = name.split('_')[0]
        userid = name.split('_')[1]

        '''
        This line retrieves the current time using datetime.now() function, which returns a datetime 
        object representing the current date and time. strftime() method is then used to format the 
        datetime object into a string with the desired format ("%H:%M:%S" represents hours:minutes:seconds).
        '''
        current_time = datetime.now().strftime("%H:%M:%S")

        '''
        This line checks if the userid string is not empty. strip() method is used to remove any leading 
        or trailing whitespace characters. If the userid is not empty, the condition evaluates to True, 
        indicating that a valid user ID is provided.
        '''
        if userid.strip():  # Check if userid is not an empty string
            try:
                '''
                These lines convert the userid string to an integer using int() function and then read 
                the attendance data for the current date from the CSV file into a DataFrame named df. 
                It checks if the userid_int is not present in the 'Roll' column of the DataFrame. If 
                the user ID is not already recorded in the attendance, the condition evaluates to True.
                '''
                userid_int = int(userid)
                df = pd.read_csv(f'Attendance/Attendance-{self.datetoday}.csv')
                if userid_int not in df['Roll'].values:
                    '''
                    this block of code opens the CSV file for appending mode and writes a new attendance 
                    record in the format "username,userid,current_time" to the file. Each record is 
                    written on a new line in the CSV file.
                    '''
                    with open(f'Attendance/Attendance-{self.datetoday}.csv', 'a') as f:
                        f.write(f'\n{username},{userid_int},{current_time}')

            except ValueError:
                '''
                If the userid cannot be converted to an integer (due to non-numeric characters), 
                a ValueError is raised. This block catches the exception and prints a message 
                indicating that the user ID provided is invalid.
                '''
                print(f"Invalid userid: {userid}")
        else:
            '''
            If the userid is empty (after stripping whitespace), this block prints a message 
            indicating that an empty user ID is provided.
            '''
            print("Empty userid provided")

    def update_students_excel(self, name, mis):
        '''
        Update the students' Excel file with a new entry.

        Parameters:
            name (str): The name of the student.
            mis (str): The Management Information System (MIS) number of the student.

        Returns:
            None
        '''

        '''
        Checks if the Excel file named 'students.xlsx' exists. 
        If it doesn't exist, it creates a new Excel file with headers 'Name' and 'MIS'.
        '''

        '''
        This line loads an Excel workbook named 'students.xlsx' into memory using the load_workbook() 
        function from the openpyxl library. The wb variable now holds a reference to this workbook.
        '''
        wb = load_workbook('students.xlsx')

        '''
        This line selects the active worksheet within the workbook loaded in the previous step. 
        The active attribute returns the currently active worksheet, and it is assigned to the 
        variable ws.
        '''
        ws = wb.active

        '''
        This line appends a new row of data to the selected worksheet (ws). 
        The data to be appended is provided as a list [name, mis]. Here, name and mis are variables. 
        This line effectively adds a new row to the end of the worksheet with the provided data.
        '''
        ws.append([name, mis])

        '''
        After appending the data, this line saves the modified workbook back to the Excel file 
        'students.xlsx'. The changes made to the worksheet (such as adding the new row of data) 
        are now persisted in the Excel file.
        '''
        wb.save('students.xlsx')    

    def getallusers(self):
        '''
        Defines a function getallusers() that retrieves information about all registered users. 

        Get all users from the 'static/faces' directory.
    
        Returns:
            tuple: A tuple containing:
                - list: List of all user filenames.
                - list: List of user names.
                - list: List of user rolls.
                - int: Number of users.
        '''

        '''
        This line uses the os.listdir() function to retrieve a list of all files and directories within 
        the 'static/faces' directory. The result is stored in the userlist variable.
        '''
        userlist = os.listdir('static/faces')

        '''
        Two empty lists, names and rolls, are initialized. These lists will be used to store the names 
        and roll numbers of the users extracted from the filenames.
        '''
        names = []
        rolls = []

        '''
        The length of the userlist is calculated using the len() function, 
        and the result is stored in the variable l.
        '''
        l = len(userlist)

        '''
        This line starts a loop that iterates over each item in the userlist. 
        In this context, each item represents a filename in the 'static/faces' directory.
        '''
        for i in userlist:
            '''
            For each filename (i), this line splits the filename using the underscore ('_') as a delimiter. 
            The resulting substrings are assigned to the variables name and roll.
            '''
            name, roll = i.split('_')
            names.append(name)
            rolls.append(roll)

        '''
        After iterating through all filenames, the function returns a tuple containing four elements: userlist 
        (the list of filenames), names (the list of extracted names), rolls (the list of extracted roll 
        numbers), and l (the length of the userlist).
        '''

        return userlist, names, rolls, l

    def mark_attendance(self):

        marker = AttendanceMarker(
            app_running=True,
            imgBackground=self.imgBackground,  # Assuming imgBackground is defined elsewhere
            extract_attendance=self.extract_attendance,
            totalreg=self.totalreg,
            datetoday2=self.datetoday2,  # Assuming datetoday2 is defined elsewhere
            extract_faces=self.extract_faces,
            identify_face=self.identify_face,
            add_attendance=self.add_attendance  # Assuming add_attendance is defined elsewhere
        )
        return marker.mark_attendance()

    def add_new_user(self):
        '''
        Route for adding a new user to the attendance system.

        Methods:
        - GET: Renders the 'newuser.html' template with attendance data.
        - POST: Processes form data to add a new user and capture their images.

        Returns:
        - GET: Rendered 'newuser.html' template.
        - POST: Rendered 'newuser.html' template with updated attendance data.
        '''

        '''
        This conditional block checks if the incoming request is a POST request. 
        If it is, it means that the form has been submitted with user data.
        '''
        if request.method == 'POST':
            # This code retrieves the username and userid values from the form submitted via POST request.
            username = request.form['username']
            userid = request.form['userid']

            # This function call updates the Excel file (students.xlsx) with the new user's information.
            self.update_students_excel(username, userid)
            
            userimagefolder = 'static/faces/' + username + '_' + str(userid)

            '''
            Checks if the directory named 'static/faces' exists. 
            If it doesn't exist, it creates the directory. This directory is used to store images of faces.
            '''
            if not os.path.isdir(userimagefolder):
                os.makedirs(userimagefolder)

            # This line initializes two variables i and j to keep track of the number of images captured (i) and 
            # the total number of iterations (j) in the loop.
            i, j = 0, 0

            # This line opens the default webcam (index 0) for video capture using the OpenCV VideoCapture object.
            cap = cv2.VideoCapture(0)
            
            # This initiates an infinite loop to continuously capture frames from the webcam until a break condition is met.
            while 1:
                # This line reads a frame from the webcam using the cap.read() function. The underscore _ is used to 
                # discard the return value indicating whether the frame was successfully read.
                _, frame = cap.read()

                # This line extracts faces from the captured frame using the extract_faces() function. 
                # It returns a list of tuples containing coordinates of detected faces.
                faces = self.extract_faces(frame)
                print("Detected faces:", faces)

                # This loop iterates over each detected face, where (x, y) represents the top-left corner coordinates of 
                # the bounding box, and (w, h) represents the width and height of the bounding box.
                for (x, y, w, h) in faces:
                    '''
                    This line draws a rectangle around the detected face on the frame. 
                    The rectangle is drawn using the cv2.rectangle() function with the specified 
                    color (255, 0, 20) and thickness 2.
                    '''
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 20), 2)

                    # This line adds text to the frame indicating the number of images captured (i) 
                    # out of the total number of images to capture (nimgs).
                    cv2.putText(frame, f'Images Captured: {i}/{self._nimgs}', (30, 30),
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 20), 2, cv2.LINE_AA)
                    
                    '''
                    This block saves every 5th detected face as an image in the user's image folder. 
                    It constructs a filename based on the username, i (image index), and saves the 
                    cropped face region as a JPEG image.
                    '''
                    if j % 5 == 0:
                        name = username + '_' + str(i) + '.jpg'
                        cv2.imwrite(userimagefolder + '/' + name, frame[y:y + h, x:x + w])
                        i += 1
                    j += 1
                
                # This condition breaks the loop once the 30 images (nimgs) have been captured.
                if j == self._nimgs * 5:
                    break

                # This line displays the frame with detected faces in a window titled 'Adding new User' using cv2.imshow().
                cv2.imshow('Adding new User', frame)

                # his line checks for the press of the 'Esc' key (ASCII code 27). 
                # If the key is pressed, it breaks out of the loop, terminating the image capture process.
                if cv2.waitKey(1) == 27:
                    break

            # releases the webcam resource using cap.release(), allowing other applications to access the webcam.
            cap.release()

            # closes all OpenCV windows created during the image capture process.
            cv2.destroyAllWindows()

            # prints a message indicating the start of the training process and then triggers the 
            # training of the face recognition model using the captured images.
            print('Training Model')
            self.train_model()
            names, rolls, times, l = self.extract_attendance()
            return render_template('newuser.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.totalreg(), datetoday2=self.datetoday2)
            
        else:
            names, rolls, times, l = self.extract_attendance()
            return render_template('newuser.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.totalreg(), datetoday2=self.datetoday2)
        








import mysql.connector
import csv

class MySQLDataExtractor:
    def __init__(self, host, user, password, database):
        self.host = host
        self.user = user
        self.password = password
        self.database = database
        self.conn = None

    def connect(self):
        try:
            self.conn = mysql.connector.connect(
                host=self.host,
                user=self.user,
                password=self.password,
                database=self.database
            )
            if self.conn.is_connected():
                print('Connected to MySQL database')
        except mysql.connector.Error as e:
            print(f'Error connecting to MySQL: {e}')

    def fetch_data_to_csv(self, table_name, csv_file):
        try:
            cursor = self.conn.cursor()
            cursor.execute(f'SELECT * FROM {table_name}')
            rows = cursor.fetchall()

            with open(csv_file, 'w', newline='') as csvfile:
                csv_writer = csv.writer(csvfile)
                csv_writer.writerows(rows)

            print(f'Data written to {csv_file} successfully')

        except mysql.connector.Error as e:
            print(f'Error fetching data from MySQL: {e}')

        finally:
            if cursor:

                # This line closes the cursor object cursor.
                cursor.close()

    def close_connection(self):
        if self.conn:
            self.conn.close()
            print('Connection closed.')

class DataProcessor:
    def __init__(self, extractor):
        self.extractor = extractor

    def process_data(self, table_name, csv_file):
        self.extractor.connect()
        if self.extractor.conn:
            self.extractor.fetch_data_to_csv(table_name, csv_file)
            self.extractor.close_connection()

















class WebPageHandler:
    def __init__(self, face_recognition_system):
        self.face_recognition_system = face_recognition_system

    def new_home(self):
        '''
        Route handler for the home page.
        Defines the new_home() function, which will handle the HTTP request to the root URL.

        Retrieves attendance data and renders the 'new_home.html' template with the data.

        Returns:
            rendered_template: HTML template with attendance data.
        '''
        
        '''
        Calls the extract_attendance() function to retrieve attendance data from a CSV file. 
        This function likely reads the attendance data from a file and returns lists of names, 
        rolls, times, and the length of the attendance records.
        '''
        names, rolls, times, l = self.face_recognition_system.extract_attendance(self)

        '''
        Uses the render_template() function provided by Flask to render the 'new_home.html' template. It passes several variables to the template:
            ◦ names: List of names extracted from attendance data.
            ◦ rolls: List of rolls extracted from attendance data.
            ◦ times: List of times extracted from attendance data.
            ◦ l: Length of the attendance records.
            ◦ totalreg(): Calls the totalreg() function, which likely returns the total number of 
                            registered users.
            ◦ datetoday2: A formatted string representing the current date.
        '''
        return render_template('new_home.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.face_recognition_system.totalreg(self), datetoday2=self.face_recognition_system.datetoday2)

    def display_attendance_page(self):
        '''
        Display the attendance page.

        Retrieves attendance data and renders the attendance.html template with the data.

        Returns:
            str: Rendered HTML page displaying attendance information.
        '''
        names, rolls, times, l = self.face_recognition_system.extract_attendance(self)
        return render_template('attendance.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.face_recognition_system.totalreg(self), datetoday2=self.face_recognition_system.datetoday2)

    def display_newuser_page(self):
        names, rolls, times, l = self.face_recognition_system.extract_attendance(self)
        return render_template('newuser.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.face_recognition_system.totalreg(self), datetoday2=self.face_recognition_system.datetoday2)

    def display_qr_page(self):

        '''
        This line calls a function extract_attendance() to retrieve attendance data, which 
        presumably populates names, rolls, times, and l variables.
        '''
        names, rolls, times, l = self.face_recognition_system.extract_attendance(self)

        '''
        This line returns a rendered HTML template (qr.html) along with data to be displayed on the template. It passes 
        attendance-related data such as names, rolls, times, l, as well as some other variables to the template \
        for rendering.
        '''
        return render_template('qr.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.face_recognition_system.totalreg(self), datetoday2=self.face_recognition_system.datetoday2)

    def display_how_page(self):
        '''
        This line calls a function extract_attendance() to retrieve attendance data, which 
        presumably populates names, rolls, times, and l variables.
        '''
        names, rolls, times, l = self.face_recognition_system.extract_attendance(self)
        return render_template('howitworks.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.face_recognition_system.totalreg(self), datetoday2=self.face_recognition_system.datetoday2)


    def display_calculate_attendance_page(self):
        '''
        This line calls a function extract_attendance() to retrieve attendance data, which 
        presumably populates names, rolls, times, and l variables.
        '''
        names, rolls, times, l = self.face_recognition_system.extract_attendance(self)
        return render_template('check_percentage.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.face_recognition_system.totalreg(self), datetoday2=self.face_recognition_system.datetoday2)



class URLRouter:
    def __init__(self, app, webpage_handler):
        self.app = app
        self.webpage_handler = webpage_handler

    def add_url_rules(self):
        self.app.add_url_rule('/', 'new_home', self.webpage_handler.new_home)
        self.app.add_url_rule('/attendance_page', 'display_attendance_page', self.webpage_handler.display_attendance_page)
        self.app.add_url_rule('/newuser_page', 'display_newuser_page', self.webpage_handler.display_newuser_page)
        self.app.add_url_rule('/qr_page', 'display_qr_page', self.webpage_handler.display_qr_page)
        self.app.add_url_rule('/how_it_works_page', 'display_how_page', self.webpage_handler.display_how_page)
        self.app.add_url_rule('/check_percentage', 'display_calculate_attendance_page', self.webpage_handler.display_calculate_attendance_page)
        





class WebInterface(URLRouter):
    def __init__(self, face_recognition_system,AttendanceManager):

        # Call the constructor of the superclass (URLRouter)
        super().__init__(app=None, webpage_handler=None)

        self.face_recognition_system = face_recognition_system
        self.AttendanceManager = AttendanceManager



        
        self.app = Flask(__name__)

        '''
        Initializes a boolean variable app_running and sets it to True. 
        This variable is used later to control the running state of the Flask application.
        '''
        self.app_running = True


        self.webpage_handler = WebPageHandler(face_recognition_system=self.face_recognition_system)
        self.url_router = URLRouter(app=self.app, webpage_handler=self.webpage_handler)
        self.url_router.add_url_rules()

        self.app.add_url_rule('/attendance', 'attendance', self.attendance, methods=['GET'])
        self.app.add_url_rule('/newuser', 'newuser', self.newuser, methods=['GET', 'POST'])
        self.app.add_url_rule('/qr_attendance', 'process_qr_attendance', self.process_qr_attendance, methods=['GET'])

        pass    

    
    def attendance(self):
        # Call the attendance method of the FaceRecognitionSystem
        return self.AttendanceManager.mark_attendance(self)
    
    def newuser(self):
        return self.AttendanceManager.add_new_user(self)
    
    def process_qr_attendance(self):
        class MySQLConnector:
            def init(self, host, user, password, database):
                self.host = host
                self.user = user
                self.password = password
                self.database = database
                self.conn = None

            def connect(self):
                try:
                    self.conn = mysql.connector.connect(
                        host=self.host,
                        user=self.user,
                        password=self.password,
                        database=self.database
                    )
                    print("Connected to MySQL database successfully!")
                except mysql.connector.Error as e:
                    print(f"Error connecting to MySQL database: {e}")

            def close_connection(self):
                if self.conn:
                    self.conn.close()
                    print("MySQL connection closed.")

        # Example usage:
        if __name__ == "main":
            # Replace these values with your MySQL connection details
            host = 'localhost'
            user = 'arnav-rppoop1'
            password = 'Guitar@123'
            database = 'db_1'

            # Create an instance of MySQLConnector
            mysql_connector = MySQLConnector(host, user, password, database)

            # Connect to MySQL
            mysql_connector.connect()

            # Perform database operations...

            # Close the connection when done
            mysql_connector.close_connection()


        # GET DATA FROM LOGIN CSV TO RESULT TABLE WITH INITIAL COUNT=1
        import pandas as pd

        # Path to your Excel file
        excel_file_path = 'students.xlsx'

        # Read the Excel file into a DataFrame
        df_login = pd.read_excel(excel_file_path)

        # Print the DataFrame
        df_login.rename(columns={'Name':'NAME'},inplace=True)
        print(df_login)

        import pandas as pd
        import mysql.connector

        # Replace these values with your MySQL connection details
        host = 'localhost'
        user = 'arnav-rppoop1'
        password = 'Guitar@123'
        database = 'db_1'

        # Connect to MySQL
        conn = mysql.connector.connect(host=host, user=user, password=password, database=database)

        query = "SELECT * FROM result"



        cursor = conn.cursor()
        # cursor.execute(query)
        # Fetch data into a DataFrame
        df_sql = pd.read_sql(query, conn)

        print(df_sql)


        # Iterate over the rows using iloc
        for i in range(len(df_login)):
            row = df_login.iloc[i]
            print("Row", i, ":", row['NAME'], row['MIS'])
            # Row index and attribute to check
            row_index = i  # Index of the row to check
            attribute_to_check = 'NAME'  # Attribute to check in df1

            # Get the value of the attribute to check
            value_to_check = df_login.loc[row_index, attribute_to_check]
            # Check if the row is present in df2
            # Check if the value is present in df2
            value_present_in_df2 = value_to_check in df_sql[attribute_to_check].values

            if not(value_present_in_df2):
                print(f"The value {value_to_check} from attribute {attribute_to_check} is not present in df2")
                print(type(df_login.loc[i,'NAME']))
                print(type(df_login.loc[i,'MIS']))
                select_query="INSERT INTO result VALUES ('" + df_login.loc[i,'NAME'] + "'," + str(df_login.loc[i,'MIS']).strip() + "," + "1)"
                print(select_query)
                # exit(0)
                cursor = conn.cursor()
                cursor.execute(select_query)    

        # Close the connection
        print('done!!')
        conn.commit()
        conn.close()
            

        # GET DATA FROM QR CSV AND FINALLY PUTTING IT IN GGG.CSV FILE(QR_DATA.CSV)


        from datetime import datetime

        import pandas as pd
        import csv

        from datetime import datetime, timedelta

        # URL of the CSV file
        url='https://docs.google.com/spreadsheets/d/e/2PACX-1vQX5dM1IhQjnTaUhy3Km-leCyvtaSqz80MRvpG6FCXSk7F9hXugsOJe44GVrH8EjRVJHTWpr0giJ1WH/pub?output=csv'
        # Read the CSV file into a pandas DataFrame
        data = pd.read_csv(url)

        # file_name='abc.csv'


        ff=[]

        c=0

        for i in data.LatLong:
            if ( type(i) != float):   
                ff.append(data.iloc[c])
            c = c+1
            
        gt = pd.DataFrame(ff)
        print(gt)




        # Get current date and time
        current_time = datetime.now()


        # x1 = current_time.year
        # x2= current_time.month
        # x3 = current_time.day
        # x4 = current_time.second
        x5 = current_time.minute
        # x6 = current_time.hour
        # x7 = current_time.microsecond

        # ddd.csv
        csv_file_path1 = 'qr_csv_data.csv'

        pp=[]
        # dataframe : gt
        # Iterate over the index and print each row
        for index in range(len(gt)):
            # print(type(gt.iloc[index]['LatLong']))
            # print(gt.iloc[index]['LatLong'])
            # Split the string based on comma
            parts = gt.iloc[index]['LatLong'].split(',')

            # Extract the numbers and convert them to floats
            latitude = float(parts[0])
            longitude = float(parts[1])
            # print(latitude)
            # 18.5300277778
            # 18.5319444444444444
            if (latitude >= 0 and  latitude<=100):
                # 73.854
                # 73.85805555555556
                if (longitude >= 0 and  longitude <=100):
                    pp.append(gt.iloc[index])

        yt = pd.DataFrame(pp)
        # ddd.csv
        csv_file_path1 = 'qr_csv_data.csv'

        # Write DataFrame to CSV, overwriting existing content
        yt.to_csv(csv_file_path1, index=False, mode='w')

        # x1 = current_time.year
        # x2= current_time.month
        # x3 = current_time.day
        # x4 = current_time.second
        x5 = current_time.minute
        # x6 = current_time.hour
        # x7 = current_time.microsecond

        # print(x5)
        # exit(0)


        cc=[]
        minutes_to_subtract = 20
        # yt
        for index in range(len(yt)):
            date_string = yt.iloc[index]['Timestamp']

            date_object = datetime.strptime(date_string, "%m/%d/%Y %H:%M:%S")
            print(date_string)
            print(type(date_string))
            print(date_object)
            print(type(date_object))
            print(current_time)
            print(type(current_time))

            time_difference = current_time-date_object

            print(time_difference)
            print(time_difference.days)
            print(type(time_difference.days))
            print(time_difference.seconds)
            print(time_difference.seconds/60)
            # exit(0)

            # if (time_difference <= timedelta(minutes=30)):
            # if time_difference.days == 0 and (int(time_difference.seconds / 60) >= 0 or int(time_difference.seconds / 60) <= 30):
            # if 1:
            if time_difference.days==0 and int(time_difference.seconds/60) <=30:
                # exit(0)
                cc.append(yt.iloc[index])
                print(cc)
                print(current_time)
                print(date_object)
                # exit(0)
        # exit(0)
                    
                
                # if (minute_1 <= x5_1 and x5<=minute_1):

                
                

            

        # 18°31'48.1"N 73°51'27.4"E
        # Latitude: 18.5300277778 N 
        # Longitude: 73.854 E
            

        # 2nd loc
        # Latitude: 18.5319444444444444∘N
        # Longitude: 73.85805555555556∘ E

        # print(gt)
                
        print(yt)
        print(x5)

        tt= pd.DataFrame(cc)
        print(tt)
        # CSV file path
        csv_file_path2 = 'qr_data.csv'

        # Write DataFrame to CSV, overwriting existing content
        tt.to_csv(csv_file_path2, index=False, mode='w')

   
        # ------------------------------------------
        # ------------------------------------
        # ------------------------------------------
        # ------------------------------------
        # ------------------------------------------
        # ------------------------------------

        # READING CSV FILE IE GGG.CSV(QR_DATA.CSV) AND PUTTING IT IN TABLE1

        import pandas as pd
        import mysql.connector

        # Assuming you already have a MySQL connection
        # Modify the connection parameters according to your database configuration
        conn = mysql.connector.connect(
            host="localhost",
            user="arnav-rppoop1",
            password="Guitar@123",
            database="db_1"
        )

        # Path to your CSV file
        file_path = 'qr_data.csv'

        # Read the CSV file into a DataFrame
        # df = pd.read_csv(file_path)
        try:
            # Read the CSV file into a DataFrame
            df = pd.read_csv(file_path)

            # Check if the DataFrame is empty
            if df.empty:
                print("The DataFrame is empty. Exiting the program.")
                exit(0)
            else:
                # Your further processing code here
                pass

        except pd.errors.EmptyDataError:
            print("The CSV file is empty or does not contain any data in the expected format. Exiting the program.")
            
            '''
            This line calls a function extract_attendance() to retrieve attendance data, which 
            presumably populates names, rolls, times, and l variables.
            '''
            names, rolls, times, l = self.extract_attendance()

            '''
            This line returns a rendered HTML template (qr.html) along with data to be displayed on the template. It passes 
            attendance-related data such as names, rolls, times, l, as well as some other variables to the template \
            for rendering.
            '''
            return render_template('qr.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.totalreg(), datetoday2=self.datetoday2)
            exit(0)

        except FileNotFoundError:
            print(f"The file '{file_path}' does not exist. Please provide a valid file path. Exiting the program.")
            
            '''
            This line calls a function extract_attendance() to retrieve attendance data, which 
            presumably populates names, rolls, times, and l variables.
            '''
            names, rolls, times, l = self.extract_attendance()

            '''
            This line returns a rendered HTML template (qr.html) along with data to be displayed on the template. It passes 
            attendance-related data such as names, rolls, times, l, as well as some other variables to the template \
            for rendering.
            '''
            return render_template('qr.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.totalreg(), datetoday2=self.datetoday2)
            exit(0)

        except Exception as e:
            print(f"An error occurred while reading the CSV file: {e}")

            '''
            This line calls a function extract_attendance() to retrieve attendance data, which 
            presumably populates names, rolls, times, and l variables.
            '''
            names, rolls, times, l = self.extract_attendance()

            '''
            This line returns a rendered HTML template (qr.html) along with data to be displayed on the template. It passes 
            attendance-related data such as names, rolls, times, l, as well as some other variables to the template \
            for rendering.
            '''
            return render_template('qr.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.totalreg(), datetoday2=self.datetoday2)
            exit(0)




        # Delete existing data from the table
        cursor = conn.cursor()
        delete_query = "DELETE FROM table1"
        cursor.execute(delete_query)

        selected_columns = ['Name', 'MIS']

        table_name = 'table1'
        # Insert new data into the MySQL table
        for index, row in df.iterrows():

            insert_query = "INSERT INTO table1 (NAME, MIS) VALUES (\'"+ row['Name'].strip() +"\',"+ str(row['MIS']) +");"
            # print(insert_query)
            # exit(0)
            cursor.execute(insert_query)

            # values = tuple(row[selected_columns])  # Select only desired columns
            # print(values)

            # query = f"INSERT INTO {table_name} (column1, column2) VALUES (%s, %s)"  # Adjust column names
            # cursor.execute(query, values)

        # Commit the changes and close the connection
        conn.commit()
        conn.close()

        # --------------------------------
        # --------------------------------
        # --------------------------------
        # --------------------------------
        # --------------------------------
        # ------------INTERSECT DATA FROM OPENCV AND QR EXCEL------------------

        import pandas as pd

        # Specify the path to your Excel file
        from datetime import datetime

        import mysql.connector

        # Connect to the MySQL database
        # Replace 'your_host', 'your_username', 'your_password', and 'your_database' with your MySQL configuration
        connection = mysql.connector.connect(
            host="localhost",
            user="arnav-rppoop1",
            password="Guitar@123",
            database="db_1"
        )

        # Create a cursor object to interact with the database
        cursor = connection.cursor()

        # Get the current time
        current_time = datetime.now()
        current_time=str(current_time)
        # Print the current time
        month=current_time[5:7]
        day=current_time[8:10]
        print("Current time:", current_time)
        print(month)
        print(day)
        print(type(month))

        excel_file = './Attendance/Attendance-'+month+"_"+day+"_24.csv"
        print(excel_file)

        # Example query
        query = "DELETE FROM table2"

        # Execute the query
        cursor.execute(query)

        # Commit the changes
        connection.commit()

        # Read the Excel file into a pandas DataFrame
        df = pd.read_csv(excel_file)

        # Display the first few rows of the DataFrame
        print(df.head())

        # Assuming df is your DataFrame
        for index, row in df.iterrows():
            # Access row data using row['column_name'] or row[index]
            # print(index, row['Column1'], row['Column2'])
            df_name = row['Name']
            df_roll=row['Roll']
            query="INSERT INTO table2 VALUES('"+str(df_name) + "'," + str(df_roll) +")"
            print(query)
            cursor.execute(query)
        # Commit the changes
        connection.commit()

        # Example query
        query = "DELETE FROM save_res"

        # Execute the query
        cursor.execute(query)

        # Commit the changes
        connection.commit()

        # query = "INSERT INTO SAVE_RES(NAME,MIS)((SELECT NAME,MIS FROM TABLE1)INTERSECT(SELECT NAME,MIS FROM TABLE2))"
        query="""INSERT INTO save_res (NAME, MIS)
        SELECT NAME, MIS FROM table1
        INTERSECT
        SELECT NAME, MIS FROM table2;
        """
        cursor.execute(query)

        # Commit the changes
        connection.commit()

        # Close the cursor and connection

        # This line closes the cursor object cursor.
        cursor.close()

        '''
        This line closes the connection to the database using the close() method of the database 
        connection object connection.
        '''
        connection.close()


        # --------------------------------------------
        # --------------------------------------------
        # --------------------------------------------
        # --------------------------------------------
        # --------------------------------------------
        # ---INCREMENT TOTAL ATTENDANCE FROM TOTAL_ATTENDANCE TABLE
        import mysql.connector

        # Connect to the MySQL database
        connection = mysql.connector.connect(
            host="localhost",
            user="arnav-rppoop1",
            password="Guitar@123",
            database="db_1"
        )

        # Create a cursor object to interact with the database
        cursor = connection.cursor()

        # Prepare the SQL query to update data
        query = """
            UPDATE total_attendance
            SET total = total + 1
        """

        # Execute the query
        cursor.execute(query)

        # Commit the changes
        connection.commit()

        # Close the cursor and connection

        # This line closes the cursor object cursor.
        cursor.close()

        '''
        This line closes the connection to the database using the close() method of the database 
        connection object connection.
        '''
        connection.close()


        # -----------------------------------
        # -----------------------------------
        # -----------------------------------
        # -----------------------------------
        # -----------------------------------
        # -------DATA FROM SAVE_RES TO RESULT TABLE


        # --------------------------------
        # --------------------------------
        # --------------------------------
        # --------------------------------
        # --PUT DATA FROM SAVE_RES TO RESULT TABLE INCREMENT COUNT ATTRIBUTE---------

        import pandas as pd
        import mysql.connector

        # Replace these values with your MySQL connection details
        host = 'localhost'
        user = 'arnav-rppoop1'
        password = 'Guitar@123'
        database = 'db_1'

        # Connect to MySQL
        conn = mysql.connector.connect(host=host, user=user, password=password, database=database)

        query = "SELECT * FROM save_res"

        cursor = conn.cursor()
        # cursor.execute(query)
        # Fetch data into a DataFrame
        df_save_res = pd.read_sql(query, conn)

        print(df_save_res)

        query = "SELECT * FROM result"

        cursor = conn.cursor()
        # cursor.execute(query)
        # Fetch data into a DataFrame
        df_final = pd.read_sql(query, conn)

        print(df_final)


        # Iterate over the rows using iloc
        for i in range(len(df_save_res)):
            row = df_save_res.iloc[i]
            print("Row", i, ":", row['NAME'], row['MIS'])
            # Row index and attribute to check
            row_index = i  # Index of the row to check
            attribute_to_check = 'NAME'  # Attribute to check in df1

            # Get the value of the attribute to check
            value_to_check = df_save_res.loc[row_index, attribute_to_check]
            # Check if the row is present in df2
            # Check if the value is present in df2
            value_present_in_df2 = value_to_check in df_final[attribute_to_check].values

            if value_present_in_df2:
                print(f"The value {value_to_check} from attribute {attribute_to_check} is not present in df2")
                print(type(df_save_res.loc[i,'NAME']))
                print(type(df_save_res.loc[i,'MIS']))
                # select_query="INSERT INTO result VALUES ('" + df_login.loc[i,'NAME'] + "'," + str(df_login.loc[i,'MIS']).strip() + "," + "1)"
                query = """
                    UPDATE result
                    SET COUNT = COUNT+1 
                    WHERE MIS =""" + str(df_save_res.loc[i,'MIS'])
                print(query)
                # exit(0)
                cursor = conn.cursor()
                cursor.execute(query) 
            else:
                print(f"The value {value_to_check} from attribute {attribute_to_check} is not present in df2")
                print(type(df_save_res.loc[i,'NAME']))
                print(type(df_save_res.loc[i,'MIS']))
                # select_query="INSERT INTO result VALUES ('" + df_login.loc[i,'NAME'] + "'," + str(df_login.loc[i,'MIS']).strip() + "," + "1)"
                query = " INSERT INTO RESULT VALUES('"+ row['NAME'] + "',"+str(row['MIS']) + "1)" 
                print(query)
                # exit(0)
                cursor = conn.cursor()
                cursor.execute(query)    

        # Close the connection
        print('done!!')
        conn.commit()
        conn.close()

        # ---------------------------------------------------
        # ---------------------------------------------------
        # ---------------------------------------------------
        # get data from database table to excel

        # get data from database table to excel
        import mysql.connector
        import pandas as pd

        # Connect to the MySQL database
        connection = mysql.connector.connect(
            host="localhost",
            user="arnav-rppoop1",
            password="Guitar@123",
            database="db_1"
        )


        # Create a cursor object to interact with the database
        cursor = connection.cursor()

        # Execute a SQL query to retrieve data
        query = "SELECT * FROM result"
        cursor.execute(query)

        # Fetch all rows from the result set
        data = cursor.fetchall()

        # Get column names from cursor description
        columns = [column[0] for column in cursor.description]

        # Create a pandas DataFrame from the fetched data
        df = pd.DataFrame(data, columns=columns)

        # Save the DataFrame to an Excel file
        # excel_file = "output.csv"

        # Specify the path where you want to save the Excel file
        excel_file = "./static/output.csv"
        # df.to_csv(excel_file)

        # Save the DataFrame to an Excel file at the specified path
        df.to_csv(excel_file, index=False)  # Set index=False to exclude row indices from the output

        # Close the cursor and connection


        # This line closes the cursor object cursor.
        cursor.close()

        # This line closes the connection to the database using the close() method of the database connection object connection.
        connection.close()


        '''
        This line prints a message indicating that the data has been successfully exported to the CSV file.
        '''
        print("Data exported to", excel_file)

        '''
        This line calls a function extract_attendance() to retrieve attendance data, which 
        presumably populates names, rolls, times, and l variables.
        '''
        names, rolls, times, l = self.extract_attendance()


        
        # ------------------------TOTAL COUNT DATA TO CSV---------------------------------------------------

        # exec("main()")
        extractor = MySQLDataExtractor(
        host='localhost',
        user='arnav-rppoop1',
        password='Guitar@123',
        database='db_1'
    )
        processor = DataProcessor(extractor)
        # # Specify the path where you want to save the Excel file
        # excel_file = "../Static/output.csv"
        # df.to_csv(excel_file)

        # Save the DataFrame to an Excel file at the specified path
        df.to_csv(excel_file, index=False)  # Set index=False to exclude row indices from the output
        processor.process_data(table_name='total_attendance', csv_file='./static/total.csv')

        '''
        This line returns a rendered HTML template (qr.html) along with data to be displayed on the template. It passes 
        attendance-related data such as names, rolls, times, l, as well as some other variables to the template \
        for rendering.
        '''
        return render_template('qr.html', names=names, rolls=rolls, times=times, l=l, totalreg=self.totalreg(), datetoday2=self.datetoday2)


    



class FaceRecognitionSystem(FlaskAppSetup, ImageProcessing, FolderFileSetup,AttendanceManager,WebInterface):
    
    # Current date formatted as "dd-Month-YYYY".
    datetoday2 = date.today().strftime("%d-%B-%Y")
    def __init__(self): 
        FlaskAppSetup.__init__(self)
        ImageProcessing.__init__(self)
        FolderFileSetup.__init__(self)
        AttendanceManager.__init__(self)
        self.setup_folders()
        self.setup_files()
        self.setup_routes()

        
        
        # Current date formatted as "dd-Month-YYYY".
        datetoday2 = date.today().strftime("%d-%B-%Y")
        # def __init__(self, face_recognition_system,AttendanceManager)
        self.web_interface = WebInterface(self,AttendanceManager)
        self.attendance_manager = AttendanceManager()  # Initialize the AttendanceManager instance
    
    # -----------------
    def run(self):
        self.app.run()    

    # Polymorphism: Same method name across different classes with different implementations
    def setup_routes(self):
        WebInterface.__init__(self, FaceRecognitionSystem,AttendanceManager)
        print("Setting up Flask routes for face recognition system.")

    def detect_faces(self, image):
        print("Detecting faces using OpenCV.")

    def setup_folders(self):
        FolderFileSetup.__init__(self)
        print("Setting up folders for face recognition system.")

    def setup_files(self):
        print("Setting up files for face recognition system.")


        

# Example usage:
if __name__ == "__main__":
    face_recognition_system = FaceRecognitionSystem()
    face_recognition_system.run()
